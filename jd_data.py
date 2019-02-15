from openpyxl import load_workbook


def read(filename):
    wb = load_workbook(filename)
    for
    sheets = wb.sheetnames
    print(type(sheets))
    print('sheet_names:' + sheets)


def write(filename):
    wb = load_workbook(filename)


if __name__ == '__main__':
    source_name = 'E:\\symdata_work\\jd_data\\sale\\20190210_all.xlsx'
    target_name = 'E:\\symdata_work\\jd_data\\result\\京东每日数据.xlsx'
    read(source_name)
    write(target_name)
